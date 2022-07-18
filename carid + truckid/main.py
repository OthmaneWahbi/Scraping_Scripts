from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys

import os
import sys
from configparser import ConfigParser

from openpyxl import Workbook
from openpyxl import load_workbook

def resource_path(relative_path: str) -> str:
    try:
        base_path = sys._MEIPASS

    except Exception:
        base_path = os.path.dirname(__file__)

    return os.path.join(base_path, relative_path)

class Vehicle(webdriver.Chrome):
    def __init__(self, driver_path = r"./driver", teardown = False):
        config = ConfigParser()
        config.read("config.ini")
        options = Options()
        options.binary_location = config.get("chrome", "chrome_location")
        options.add_experimental_option('excludeSwitches',['enable-logging'])
        self.driver_path = resource_path(config.get("chromedriver", "path"))
        self.teardown = teardown
        os.environ['PATH'] += self.driver_path
        super(Vehicle, self).__init__(chrome_options=options)
        self.implicitly_wait(60)
        self.maximize_window()

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.teardown:
            self.quit()

    def land_first_page(self):
        config = ConfigParser()
        config.read("config.ini")
        self.get(config.get('website','url'))

    def select_year(self):
        year_button = self.find_element(By.CSS_SELECTOR,"div[data-placeholder='Year']")
        make_button = self.find_element(By.CSS_SELECTOR, "div[data-placeholder='Make']")
        model_button = self.find_element(By.CSS_SELECTOR, "div[data-placeholder='Model']")

        year_button.click()
        years = self.find_elements(By.CSS_SELECTOR,"li[class='item ']")
        final_list = []

        for year in years:
            if year == years[0]:
                yearvalue = year.get_attribute('value')
                while yearvalue == '':
                    yearvalue = year.get_attribute('value')
                year.click()
                make_button.click()
                makers = self.find_elements(By.CSS_SELECTOR,"li[class='item ']")
                for make in makers:
                    if make == makers[0] and year == years[0]:
                        makevalue = make.get_attribute('innerHTML')
                        while makevalue == '':
                            makevalue = make_button.find_element(By.CSS_SELECTOR, "li[class='item  -active']").text
                        make.click()
                        model_button.click()
                        models = model_button.find_elements(By.TAG_NAME, "li")[1:]
                        for model in models:
                            modelvalue = model.get_attribute('innerHTML')
                            list_to_append = [yearvalue,makevalue,modelvalue]
                            print(list_to_append)
                            final_list.append(list_to_append)
                        make_button.click()
                    else:
                        make_button.send_keys(Keys.ARROW_DOWN)
                        makevalue = make_button.find_element(By.CSS_SELECTOR,"li[class='item  -active']").text
                        while makevalue == '':
                            makevalue = make_button.find_element(By.CSS_SELECTOR, "li[class='item  -active']").text
                        model_button = self.find_element(By.CSS_SELECTOR, "div[data-placeholder='Model']")
                        model_button.click()
                        models = model_button.find_elements(By.TAG_NAME, "li")[1:]
                        for model in models:
                            modelvalue = model.get_attribute('innerHTML')
                            list_to_append = [yearvalue,makevalue,modelvalue]
                            print(list_to_append)
                            final_list.append(list_to_append)
                        make_button.click()
            else:
                year_button.send_keys(Keys.ARROW_DOWN)
                yearvalue = year_button.find_element(By.CSS_SELECTOR, "li[class='item  -active']").get_attribute('value')
                while yearvalue == '':
                    yearvalue = year_button.find_element(By.CSS_SELECTOR, "li[class='item  -active']").get_attribute(
                        'value')
                make_button.click()
                makers = self.find_elements(By.CSS_SELECTOR, "li[class='item ']")
                for make in makers:
                    if make == makers[0] and year == years[0]:
                        makevalue = make.get_attribute('innerHTML')
                        while makevalue == '':
                            makevalue = make_button.find_element(By.CSS_SELECTOR, "li[class='item  -active']").text
                        make.click()
                        model_button.click()
                        models = model_button.find_elements(By.TAG_NAME, "li")[1:]
                        for model in models:
                            modelvalue = model.get_attribute('innerHTML')
                            list_to_append = [yearvalue, makevalue, modelvalue]
                            print(list_to_append)
                            final_list.append(list_to_append)
                        make_button.click()
                    else:
                        make_button.send_keys(Keys.ARROW_DOWN)
                        makevalue = make_button.find_element(By.CSS_SELECTOR, "li[class='item  -active']").text
                        while makevalue == '':
                            makevalue = make_button.find_element(By.CSS_SELECTOR, "li[class='item  -active']").text
                        model_button = self.find_element(By.CSS_SELECTOR, "div[data-placeholder='Model']")
                        model_button.click()
                        models = model_button.find_elements(By.TAG_NAME, "li")[1:]
                        for model in models:
                            modelvalue = model.get_attribute('innerHTML')
                            list_to_append = [yearvalue, makevalue, modelvalue]
                            print(list_to_append)
                            final_list.append(list_to_append)
                        make_button.click()
        return final_list


with Vehicle() as bot:
    config = ConfigParser()
    config.read("config.ini")
    workbook_name = '../Final_results_of_scraping.xlsx'
    try:
        wb = load_workbook(workbook_name)
    except FileNotFoundError:
        wb = Workbook()
    page = wb.create_sheet()
    if "truckid" in config.get('website','url'):
        page.title = "truckid.com"
        type = "Semi"
    elif "carid" in config.get('website','url'):
        page.title = "truckid.com"
        type = "Automobile"
    else:
        print("The website in config.ini is neither carid.com or truckid.com\nChange the website to one those two.")
        exit(0)
    bot.land_first_page()
    result = bot.select_year()
    headers = ['Type','Year','Make','Model']
    page.append(headers)
    for row in result:
        row.insert(0,"Semi")
        page.append(row)
    wb.save(filename='../Final_results_of_scraping.xlsx')
