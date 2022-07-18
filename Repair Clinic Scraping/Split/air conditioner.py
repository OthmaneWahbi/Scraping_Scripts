from selenium.webdriver.common.by import By
import undetected_chromedriver as uc
import time
from openpyxl import Workbook
from openpyxl import load_workbook

if __name__ == '__main__':
    uc.TARGET_VERSION = 104
    driver = uc.Chrome()
    driver.implicitly_wait(30)
    driver.get('https://www.repairclinic.com/Shop-For-Parts/a16/Air-Conditioner-Parts')

    type_text = "Air Conditionner"

    final_final_list = []

    brand_names_container = driver.find_element(By.CSS_SELECTOR, "div[id='linklist_facet_brand_names']")
    brands = brand_names_container.find_elements(By.TAG_NAME, 'a')
    makes = []
    links_make = []
    for brand in brands:
        links_make.append(brand.get_attribute('href'))
        makes.append(brand.text)
    for link_make, make_text in zip(links_make, makes):
        driver.get(link_make)
        print(make_text)
        popular_models = driver.find_element(By.CSS_SELECTOR, "div[id='sectionPanel_popularModels']")
        final_list = []
        try:
            driver.get(
                popular_models.find_element(By.XPATH, "//*[contains(text(), 'View all models')]").get_attribute('href'))
            test = True
            while test == True:
                models_section = driver.find_element(By.CSS_SELECTOR, "div[class='modelsSection']")
                all_models_inpage = driver.find_elements(By.CSS_SELECTOR,
                                                         "div[class='modelLink col-md-4 col-sm-6 col-xs-12']")
                for model in all_models_inpage:
                    try:
                        model = model.find_element(By.TAG_NAME, "a")
                    except:
                        pass
                for model in all_models_inpage:
                    try:
                        model_text = model.text.split()
                    except:
                        pass
                    try:
                        model_text.remove(type_text)
                    except:
                        words_to_remove = type_text.split()
                        for word in words_to_remove:
                            try:
                                model_text.remove(word)
                            except:
                                pass
                    try:
                        model_text.remove(make_text)
                    except:
                        words_to_remove = make_text.split()
                        for word in words_to_remove:
                            try:
                                model_text.remove(word)
                            except:
                                pass
                    model_final_text = ''
                    for x in model_text:
                        model_final_text = model_final_text + ' ' + x
                    list_to_append = [type_text, make_text, model_final_text]
                    if list_to_append in final_list:
                        pass
                    else:
                        final_list.append(list_to_append)
                        print(list_to_append)
                try:
                    next_button = driver.find_element(By.CSS_SELECTOR, "li[class='next']")
                    next_button = next_button.find_element(By.TAG_NAME, 'a')
                    next_button.click()
                    # driver.refresh()
                    time.sleep(1)
                except:
                    next_button = driver.find_element(By.CSS_SELECTOR, "li[class='next disabled']")
                    test = False
        except:
            list_of_models = popular_models.find_element(By.CSS_SELECTOR, "div[class='linkList']")
            models = list_of_models.find_elements(By.TAG_NAME, "a")
            for model in models:
                model_text = model.text
                list_to_append = [type_text, make_text, model_text]
                if list_to_append in final_list:
                    pass
                else:
                    final_list.append(list_to_append)
                    print(list_to_append)
        final_final_list = final_final_list + final_list

    workbook_name = 'Final_results_of_scraping.xlsx'
    try:
        load_workbook(workbook_name)
        page = wb.active
    except:
        wb = Workbook()
        headers = ['Type', 'Make', 'Model']
        page = wb.active
        page.append(headers)

    for row in final_final_list:
        page.append(row)
    wb.save(filename='Final_results_of_scraping.xlsx')